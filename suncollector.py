from pathlib import Path  # Standard Python Module
from openpyxl import load_workbook, Workbook  # pip install openpyxl
import warnings

warnings.simplefilter(action='ignore', category=UserWarning) # Remove UserWarning: Data Validation extension is not supported and will be removed ws_parser.bind_all()

wb = load_workbook(filename="test.xlsx") 
for ws in wb.worksheets:
    last_row = len(ws["A"])
    print(f"{last_row}")

